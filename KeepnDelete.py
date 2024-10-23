import openpyxl
 
# Load the workbooks
workbook1_path = r'C:\Users\Benedict.Salako\OneDrive - insidemedia.net\Desktop\keyword test.xlsx'
workbook2_path = r'C:\Users\Benedict.Salako\OneDrive - insidemedia.net\Desktop\keep and Delete.xlsx' 
output_path = r'C:\Users\Benedict.Salako\OneDrive - insidemedia.net\Desktop\keyword test_updated.xlsx'
 
workbook1 = openpyxl.load_workbook(workbook1_path)
workbook2 = openpyxl.load_workbook(workbook2_path)
 
# Get the sheets
sheet1 = workbook1["Sheet1"]
sheet2_words = workbook2["Delete"]
sheet2_exceptions = workbook2["Keep"]
 
# Extract words to delete and words to keep
words_to_delete = set([cell.value.lower() for row in sheet2_words.iter_rows() for cell in row if cell.value is not None])
words_to_keep = set([cell.value.lower() for row in sheet2_exceptions.iter_rows() for cell in row if cell.value is not None])
 
# Iterate through rows in reverse order to safely delete
for row_index in range(sheet1.max_row, 0, -1):
    row = sheet1[row_index]
    row_text = " ".join([str(cell.value).lower() if cell.value is not None else "" for cell in row])
 
    # Check if any word to delete is present and no word to keep is present
    if any(word in row_text for word in words_to_delete) and not any(word in row_text for word in words_to_keep):
        sheet1.delete_rows(row_index)
 
# Save the modified workbook to the specified output path
workbook1.save(output_path)
print(f"Rows deleted successfully! Modified workbook saved to: {output_path}")