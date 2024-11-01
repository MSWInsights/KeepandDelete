import streamlit as st
import openpyxl
import io
import os
import time

# Streamlit app title
st.title("Excel Row Deletion App")

# Upload workbook1 (the main workbook)
workbook1_file = st.file_uploader("Upload the main workbook (keyword test)", type=['xlsx'])

# Upload workbook2 (the workbook with words to delete and keep)
workbook2_file = st.file_uploader("Upload the workbook with words to delete and keep (Keep and delete)", type=['xlsx'])

if workbook1_file and workbook2_file:
    # Extract the input file name for naming the output file
    workbook1_name = os.path.splitext(workbook1_file.name)[0]
    output_filename = f"{workbook1_name}_updated.xlsx"
    
    # Load the workbooks
    workbook1 = openpyxl.load_workbook(workbook1_file)
    workbook2 = openpyxl.load_workbook(workbook2_file)

    # Inform the user that the files were successfully uploaded
    st.success("Files uploaded successfully. Click 'Start Application' to process the data.")

    # Start Application button
    if st.button('Start Application'):
        # Progress bar
        progress_bar = st.progress(0)

        # Get the sheets
        sheet1 = workbook1["Sheet1"]
        sheet2_words = workbook2["Delete"]
        sheet2_exceptions = workbook2["Keep"]

        # Extract words to delete and words to keep
        words_to_delete = set([cell.value.lower() for row in sheet2_words.iter_rows() for cell in row if cell.value is not None])
        words_to_keep = set([cell.value.lower() for row in sheet2_exceptions.iter_rows() for cell in row if cell.value is not None])

        # Iterate through rows in reverse order to safely delete
        total_rows = sheet1.max_row
        rows_deleted = 0

        for row_index in range(total_rows, 0, -1):
            row = sheet1[row_index]
            row_text = " ".join([str(cell.value).lower() if cell.value is not None else "" for cell in row])

            # Check if any word to delete is present and no word to keep is present
            if any(word in row_text for word in words_to_delete) and not any(word in row_text for word in words_to_keep):
                sheet1.delete_rows(row_index)
                rows_deleted += 1

            # Update progress bar
            progress_percentage = int(((total_rows - row_index + 1) / total_rows) * 100)
            progress_bar.progress(progress_percentage)

            # Optional: Add a slight delay to visualize progress for small files
            time.sleep(0.01)  # Remove this line for large files if speed is important

        # Save the modified workbook to an in-memory file
        output = io.BytesIO()
        workbook1.save(output)
        output.seek(0)

        # Provide the download link for the updated file
        st.success(f"Rows deleted successfully! {rows_deleted} rows were deleted.")
        st.download_button(
            label="Download the updated workbook",
            data=output,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("Please upload both Excel files to proceed.")
